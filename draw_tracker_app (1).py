import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
from fuzzywuzzy import process
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Daily Draw Tracker", layout="wide")
st.title("📅 Daily Draw Tracker Generator")

# --- File Uploads ---
pdf_files = st.file_uploader("Upload PDF Draw Approvals", type="pdf", accept_multiple_files=True)
inspections_file = st.file_uploader("Upload Inspections File (CSV or Excel)", type=["csv", "xlsx"])
sage_file = st.file_uploader("Upload Sage Export (CSV or Excel)", type=["csv", "xlsx"])

if pdf_files and inspections_file and sage_file:
    with st.spinner("Processing files..."):

        # Load inspection file
        if inspections_file.name.endswith(".csv"):
            inspections_df = pd.read_csv(inspections_file)
        else:
            inspections_df = pd.read_excel(inspections_file, skiprows=0)

        inspections_df.columns = inspections_df.columns.str.strip()
        loan_col = [col for col in inspections_df.columns if "Loan Number" in col][0]
        address_col = [col for col in inspections_df.columns if "Address" in col][0]
        inspections_df = inspections_df[[loan_col, address_col]].dropna()
        inspections_df.columns = ["Loan ID", "Property Address"]

        # Load Sage file
        if sage_file.name.endswith(".csv"):
            sage_df = pd.read_csv(sage_file)
        else:
            sage_df = pd.read_excel(sage_file, sheet_name="Projects")

        rtl_df = sage_df[sage_df["Loan Types ID"] == "RTL"]
        sage_lookup = rtl_df[["Project ID", "Location name", "Customer name"]].drop_duplicates()

        # Extract data from PDFs
        tracker_entries = []

        for file in pdf_files:
            doc = fitz.open(stream=file.read(), filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)

            address_line = next((line for line in text.splitlines() if any(k in line.lower() for k in ["st", "ave", "road", "ln", "place", "pid"])), "").strip()
            amount = pd.to_numeric(next((word.replace("$", "").replace(",", "") for word in text.split() if "$" in word), "0"), errors='coerce')
            draw_number = next((line for line in text.splitlines() if "draw" in line.lower() and "#" in line), "Draw 1")

            match = process.extractOne(address_line, inspections_df["Property Address"])
            loan_id = inspections_df.loc[inspections_df["Property Address"] == match[0], "Loan ID"].values[0] if match else ""

            company = guarantor = ""
            if loan_id:
                match_row = sage_lookup[sage_lookup["Project ID"] == loan_id]
                if not match_row.empty:
                    company = match_row["Location name"].values[0]
                    guarantor = match_row["Customer name"].values[0]

            tracker_entries.append({
                "Loan ID": loan_id,
                "Property Address": address_line,
                "Blank Column 1": "",
                "Draw Amount": amount,
                "Location Name": company,
                "# of Draw": draw_number.strip(),
                "Blank Column 2": "",
                "Guarantor": guarantor
            })

        # Output to Excel with spacing
        wb = Workbook()
        ws = wb.active
        ws.title = "Draw Tracker"

        headers = [
            "Loan ID", "Property Address", "Blank Column 1", "Draw Amount",
            "Location Name", "# of Draw", "Blank Column 2", "Guarantor"
        ]

        for col_idx, col in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=col)

        row_idx = 2
        for entry in tracker_entries:
            for col_idx, col_name in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=entry.get(col_name, ""))
            row_idx += 2  # Insert spacer row

        from io import BytesIO
        output = BytesIO()
        wb.save(output)

        st.success("Tracker generated!")
        st.download_button(
            label="📄 Download Excel Tracker",
            data=output.getvalue(),
            file_name="Daily_Draw_Tracker.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
