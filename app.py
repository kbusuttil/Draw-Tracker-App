import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openai
import os

st.set_page_config(page_title="AI Draw Tracker", layout="wide")
st.title("🧠 AI-Powered Daily Draw Tracker")

openai_api_key = st.text_input("Enter your OpenAI API key", type="password")

pdf_files = st.file_uploader("Upload PDF Draw Approvals", type="pdf", accept_multiple_files=True)

if pdf_files and openai_api_key:
    openai.api_key = openai_api_key
    tracker_entries = []
    debug_preview = []

    for file in pdf_files:
        doc = fitz.open(stream=file.read(), filetype="pdf")
        full_text = "\n".join(page.get_text() for page in doc)

        prompt = (
            "You are a financial assistant. Given the text of a draw approval PDF, extract the following fields:\n"
            "1. Full Property Address (including city/state/zip)\n"
            "2. Approved Draw Amount (just the number, no $)\n"
            "3. Draw Number (e.g. '3rd', '1st', etc.)\n"
            "Text:\n" + full_text
        )

        try:
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that extracts financial data from PDF text."},
                    {"role": "user", "content": prompt}
                ]
            )
            result = response["choices"][0]["message"]["content"]
        except Exception as e:
            result = f"Error: {e}"

        address = amount = draw = ""
        for line in result.splitlines():
            if "address" in line.lower():
                address = line.split(":", 1)[-1].strip()
            elif "amount" in line.lower():
                amount = line.split(":", 1)[-1].strip().replace("$", "").replace(",", "")
            elif "draw" in line.lower():
                draw = line.split(":", 1)[-1].strip()

        debug_preview.append({
            "File": file.name,
            "Extracted Address": address,
            "Approved Amount": amount,
            "Draw #": draw,
            "Raw GPT Reply": result
        })

        tracker_entries.append({
            "Loan ID": "",
            "Property Address": address,
            "Blank Column 1": "",
            "Draw Amount": amount,
            "Location Name": "",
            "# of Draw": draw,
            "Blank Column 2": "",
            "Guarantor": "",
            "Notes": "AI Extracted"
        })

    st.subheader("🔍 AI Extraction Preview")
    st.dataframe(pd.DataFrame(debug_preview))

    wb = Workbook()
    ws = wb.active
    ws.title = "Draw Tracker"
    headers = [
        "Loan ID", "Property Address", "Blank Column 1", "Draw Amount",
        "Location Name", "# of Draw", "Blank Column 2", "Guarantor", "Notes"
    ]
    for col_idx, col in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=col)

    row_idx = 2
    for entry in tracker_entries:
        for col_idx, col_name in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=entry.get(col_name, ""))
        row_idx += 2

    from io import BytesIO
    output = BytesIO()
    wb.save(output)

    st.success("AI-powered tracker generated!")
    st.download_button(
        label="📄 Download Excel Tracker",
        data=output.getvalue(),
        file_name="AI_Daily_Draw_Tracker.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
