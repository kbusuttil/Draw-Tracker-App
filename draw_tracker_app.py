import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
from fuzzywuzzy import process
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Daily Draw Tracker", layout="wide")
st.title("📅 Daily Draw Tracker Generator")

# --- File Uploads ---
pdf_files = st.file_uploader("Upload PDF Draw Approvals", type="pdf", accept_multiple_files=True)
inspections_file = st.file_uploader("Upload Inspections Excel (Loan ID lookup)", type=["xlsx"])
sage_file = st.file_uploader("Upload Sage Export (Company/Guarantor lookup)", type=["xlsx"])

if pdf_files and inspections_file and sage_file:
    with st.spinner("Processing files..."):

        # --- Load Inspection and Sage Data ---
        inspections_df = pd.read_excel(inspections_file, skiprows=10)  # adjust skiprows based on format
        inspections_df = inspections_df[["Loan Number - DFS  ↑", "Property 1 Address  ↑"]].dropna()
        inspections_df.columns = ["Loan ID", "Property Address"]

        sage_df = pd.read_excel(sage_file, sheet_name="Projects")
        rtl_df = sage_df[sage_df["Loan Types ID"] == "RTL"]
        sage_lookup = rtl_df[["Project ID", "Location name", "Customer name"]].drop_duplicates()

        # --- Process Each PDF ---
        tracker_entries = []

        for file in pdf_files:
            doc = fitz.open(stream=file.read(), filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)

            # --- Extract Address ---
            address_line = next((line for line in text.splitlines() if any(k in line.lower() for k in ["st", "ave", "road", "ln", "place", "pid"])), "").strip()
            amount_match = pd.to_numeric(next((word.replace("$", "").replace(",", "") for word in text.split() if "$" in word), "0"), errors='coerce')
            draw_match = next((line for line in text.splitlines() if "draw" in line.lower() and "#" in line), "Draw 1")

            # --- Fuzzy Match to Inspection Data ---
            match = process.extractOne(address_line, inspections_df["Property Address"])
            loan_id = inspections_df.loc[inspections_df["Property Address"] == match[0], "Loan ID"].values[0] if match else ""

            # --- Match to Sage ---
            company = guarantor = ""
            if loan_id:
                match_row = sage_lookup[sage_lookup["Project ID"] == loan_id]
                if not match_row.empty:
                    company = match_row["Location name"].values[0]
                    guarantor = match_row["Customer name"].values[0]

            tracker_entries.append({
                "Loan ID": loan_id,
                "Property Address": address_line,
                "Blank Column 1": "",
                "Draw Amount": amount_match,
                "Location Name": company,
                "# of Draw": draw_match.strip(),
                "Blank Column 2": "",
                "Guarantor": guarantor
            })

        # --- Export to Excel ---
        df_out = pd.DataFrame(tracker_entries)
        wb = Workbook()
        ws = wb.active
        ws.title = "Draw Tracker"

        headers = df_out.columns.tolist()
        for col_idx, col in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=col)

        row_idx = 2
        for _, row in df_out.iterrows():
            for col_idx, col_name in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            row_idx += 2  # spacer row

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        st.success("Tracker generated!")
        st.download_button("📄 Download Excel Tracker", data=output.getvalue(), file_name="Daily_Draw_Tracker.xlsx")
