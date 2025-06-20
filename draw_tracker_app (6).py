import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
from fuzzywuzzy import process, fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

st.set_page_config(page_title="Daily Draw Tracker", layout="wide")
st.title("📅 Daily Draw Tracker Generator")

# --- File Uploads ---
pdf_files = st.file_uploader("Upload PDF Draw Approvals", type="pdf", accept_multiple_files=True)
inspections_file = st.file_uploader("Upload Inspections File (CSV or Excel)", type=["csv", "xlsx"])
sage_file = st.file_uploader("Upload Sage Export (CSV or Excel)", type=["csv", "xlsx"])

if pdf_files and inspections_file and sage_file:
    with st.spinner("Processing files..."):

        # Load inspection file
        if inspections_file.name.endswith(".csv"):
            inspections_df = pd.read_csv(inspections_file)
        else:
            inspections_df = pd.read_excel(inspections_file, skiprows=0)

        inspections_df.columns = inspections_df.columns.str.strip()
        loan_col = [col for col in inspections_df.columns if "Loan Number" in col][0]
        address_col = [col for col in inspections_df.columns if "Address" in col][0]
        inspections_df = inspections_df[[loan_col, address_col]].dropna()
        inspections_df.columns = ["Loan ID", "Property Address"]

        # Load Sage file
        if sage_file.name.endswith(".csv"):
            sage_df = pd.read_csv(sage_file)
        else:
            sage_df = pd.read_excel(sage_file, sheet_name="Projects")

        rtl_df = sage_df[sage_df["Loan Types ID"] == "RTL"]
        sage_lookup = rtl_df[["Project ID", "Location name", "Customer name"]].drop_duplicates()

        # Extract data from PDFs
        tracker_entries = []
        debug_preview = []

        def extract_street_number(text):
            match = re.match(r"(\d+)", text)
            return match.group(1) if match else ""

        for file in pdf_files:
            doc = fitz.open(stream=file.read(), filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            lines = text.splitlines()

            # Extract valid address line
            address_line = next((line.strip() for line in lines if re.search(r'\d+\s+.*\b(Street|St|Avenue|Ave|Road|Rd|Lane|Ln|Place|Pl|Drive|Dr|Court|Ct|NW|SW|NE|SE)\b', line)), "")
            if not re.search(r'\d+\s+[A-Za-z]', address_line):
                address_line = ""

            # Extract Draw Amount
            amount = 0.0
            for line in lines:
                if "Approved Draw Release Amount" in line:
                    amt_match = re.search(r'\$([\d,]+\.\d{2})', line)
                    if amt_match:
                        amount = float(amt_match.group(1).replace(",", ""))
                        break

            # Extract Draw Number
            draw_number_raw = next((line for line in lines if re.search(r'Draw\s+#?\d+', line, re.IGNORECASE)), "Draw #1")
            draw_match = re.search(r'Draw\s+#?(\d+)', draw_number_raw, re.IGNORECASE)
            draw_number = "1st"
            if draw_match:
                num = int(draw_match.group(1))
                suffix = "th" if 11 <= num % 100 <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(num % 10, "th")
                draw_number = f"{num}{suffix}"

            # Fuzzy match to inspections
            loan_id = ""
            if address_line:
                input_street_num = extract_street_number(address_line)
                best_match = process.extractOne(address_line, inspections_df["Property Address"], scorer=fuzz.token_set_ratio)
                if best_match and best_match[1] >= 85:
                    match_addr = best_match[0]
                    match_street_num = extract_street_number(match_addr)
                    if input_street_num == match_street_num:
                        loan_id = inspections_df.loc[inspections_df["Property Address"] == match_addr, "Loan ID"].values[0]

            # Sage lookup
            company = guarantor = ""
            if loan_id:
                match_row = sage_lookup[sage_lookup["Project ID"] == loan_id]
                if not match_row.empty:
                    company = match_row["Location name"].values[0]
                    guarantor = match_row["Customer name"].values[0]

            # Append debug info
            debug_preview.append({
                "File Name": file.name,
                "Extracted Address": address_line,
                "Draw Amount": amount,
                "Draw #": draw_number,
                "Matched Loan ID": loan_id,
                "Company": company,
                "Guarantor": guarantor
            })

            tracker_entries.append({
                "Loan ID": loan_id,
                "Property Address": address_line,
                "Blank Column 1": "",
                "Draw Amount": amount,
                "Location Name": company,
                "# of Draw": draw_number,
                "Blank Column 2": "",
                "Guarantor": guarantor
            })

        # Preview before export
        st.subheader("🔍 Preview Extracted Draw Data")
        st.dataframe(pd.DataFrame(debug_preview))

        # Output to Excel with spacing
        wb = Workbook()
        ws = wb.active
        ws.title = "Draw Tracker"

        headers = [
            "Loan ID", "Property Address", "Blank Column 1", "Draw Amount",
            "Location Name", "# of Draw", "Blank Column 2", "Guarantor"
        ]

        for col_idx, col in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=col)

        row_idx = 2
        for entry in tracker_entries:
            for col_idx, col_name in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=entry.get(col_name, ""))
            row_idx += 2

        from io import BytesIO
        output = BytesIO()
        wb.save(output)

        st.success("Tracker generated!")
        st.download_button(
            label="📄 Download Excel Tracker",
            data=output.getvalue(),
            file_name="Daily_Draw_Tracker.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
